import openpyxl
import csv

import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt

wb_obj = openpyxl.load_workbook('../luck_generation_data_frame.xlsx')
sheet = wb_obj.active

col_names = []
candidates = []
count = [[0] * sheet.max_column for _ in range(sheet.max_row)]
luck = [[0] * sheet.max_column for _ in range(sheet.max_row)]

# Import sets of candidates
for i, row in enumerate(sheet.iter_rows(1, sheet.max_row)):
    for j, column in enumerate(sheet.iter_cols(1, sheet.max_column)):
        col_names.append(column[i].value)
    candidates.append([col_names[1], col_names[2], col_names[3]])
    col_names = []
candidates = candidates[1:]
print(candidates)

# Prepare Relevant and Surprise indices
sup_indices = sorted(list(set([x[0] for x in candidates])))
rel_indices = sorted(list(set([x[1] for x in candidates])))
res = [[0] * len(rel_indices) for _ in range(len(sup_indices))]
sup_indices = dict(zip(sup_indices, range(len(sup_indices))))
rel_indices = dict(zip(rel_indices, range(len(rel_indices))))

print(sup_indices)
print(f'R {len(rel_indices)} S {len(sup_indices)}')
# Calculate matrix
for x in candidates:
    res[sup_indices[x[0]]][rel_indices[x[1]]] = x[2]

with open('heatmap.csv', 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerows(res)

###########################################################
 # TODO : fix color range.
# df = pd.read_csv('heatmap.csv')
# heatmap = ((np.asarray(df)))
# fig, ax = plt.subplots(figsize=(60,60))
# title = "Luck Generation Heatmap"
#
# plt.title(title, fontsize=18)
# ttl = ax.title
# ttl.set_position([0.5, 1.05])
#
# ax.set_xticks([])
# ax.set_yticks([])
# ax.axis('off')
#
# sns.heatmap(heatmap, annot=heatmap, fmt="", cmap='RdYlGn', linewidths=0.30, ax=ax)
# sns.heatmap(df, cmap="YlGnBu")
# plt.show()